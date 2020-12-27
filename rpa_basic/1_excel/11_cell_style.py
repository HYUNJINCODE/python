from openpyxl import load_workbook
from openpyxl import styles
from openpyxl.styles import Font, Border, Side
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B2"]
c1 = ws["C3"]

ws.column_dimensions["A"].width =5
#A열 너비 5
ws.row_dimensions[1].height= 50
a1.font = Font(color="FF0000",italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF",size=20, underline="single")
# change color

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"))
#틀고정
ws.freeze_panes = "B2"
wb.save("sample_style.xlsx")