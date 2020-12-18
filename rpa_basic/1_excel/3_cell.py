from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title ="hyunjinsheet"
ws["A1"] = 1
#a1에 1입력

ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5

print(ws["A1"]) #셀 객체를 출력
print(ws["A1"].value) #셀 객체의 값을 출력
print(ws["A10"].value) #값없을시 None 출력
print(ws.cell(row=1,column=1).value) #row column 순서 바꾸기가능.
c= ws.cell(column =3 ,row=1, value =10) #ws["C1"].value =10
print(c.value)

from random import *
#반복문을 이용해서 랜덤 숫자 채우기
for x in range(1,11): 
    for y in range(1,11):
        ws.cell(row=x, column =y , value = randint(0,100))
wb.save("sample.xlsx")