from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active
#한줄씩 데이터 넣기
ws.append(["번호","영어","수학"]) #A,B,C
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])
col_B = ws["B"] #영어 column 만 가져온다
# print(col_B)

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]
#영어수학컬럼 함께가져온다.
for cols in col_range:
    for cell in cols:
        print(cell.value)


row_title = ws[1] #1 row 
for cell in row_title:
    print(cell.value)
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()
from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        
        print(cell.value, end=" ")
        print(cell.coordinate, end=" ")
    print()

print(tuple(ws.rows))
print(tuple(ws.columns))
    
wb.save("sample.xlsx")
