from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 sheet 기본이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" #rgb 형태로 값을 넣어주면 탭색상변경
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2)  #2번째 index에 sheet생성
new_ws = wb["NewSheet"]
print(wb.sheetnames) #모든 시트이름 확인
#sheet복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"
wb.save("sample.xlsx")

